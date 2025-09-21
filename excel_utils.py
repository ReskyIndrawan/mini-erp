import os
import datetime
import json
import sys
from pathlib import Path
from tkinter import filedialog
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

TITLE = "不具合品一覧表"
EXCEL_NAME = "不具合品一覧表.xlsx"

# History management constants
APP_NAME = "defect_data_app"
HISTORY_FILENAME = "recent_excel_files.json"
MAX_HISTORY = 10


def get_config_dir():
    """Get cross-platform config directory"""
    if sys.platform.startswith("win"):
        appdata = os.environ.get("APPDATA")
        if appdata:
            return Path(appdata) / APP_NAME
    # Prefer XDG_CONFIG_HOME if set, otherwise ~/.config
    xdg = os.environ.get("XDG_CONFIG_HOME")
    if xdg:
        return Path(xdg) / APP_NAME
    return Path.home() / ".config" / APP_NAME


class ExcelHistoryManager:
    def __init__(self, max_items=MAX_HISTORY):
        self.max_items = max_items
        self.config_dir = get_config_dir()
        self.config_dir.mkdir(parents=True, exist_ok=True)
        self.path = self.config_dir / HISTORY_FILENAME
        self._items = self._load()

    def _load(self):
        """Load history from JSON file"""
        if self.path.exists():
            try:
                with open(self.path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                # ensure it's a list of strings
                if isinstance(data, list):
                    return [str(x) for x in data if isinstance(x, str)]
            except Exception:
                pass
        return []

    def save(self):
        """Save history to JSON file"""
        try:
            with open(self.path, "w", encoding="utf-8") as f:
                json.dump(self._items, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print("Failed saving recent files:", e)

    def add(self, filepath):
        """Add file to history"""
        filepath = str(Path(filepath).resolve())
        if filepath in self._items:
            self._items.remove(filepath)
        self._items.insert(0, filepath)
        self._items = self._items[: self.max_items]
        self.save()

    def clear(self):
        """Clear all history"""
        self._items = []
        try:
            if self.path.exists():
                self.path.unlink()
        except Exception:
            pass

    def items(self):
        """Get all history items"""
        return list(self._items)

    def remove(self, filepath):
        """Remove specific file from history"""
        f = str(Path(filepath).resolve())
        if f in self._items:
            self._items.remove(f)
            self.save()


def create_excel_if_not_exists(folder, creator):
    filepath = os.path.join(folder, EXCEL_NAME)
    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Title
        ws.merge_cells("A1:L1")
        ws["A1"] = TITLE
        ws["A1"].font = Font(size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Date & creator
        today = datetime.date.today()
        ws["A2"] = f"月間期間: {today.year}-{today.month}"
        ws["C2"] = f"作成者: {creator}"

        # Header
        headers = [
            "発生月",
            "累計",
            "№",
            "発生日",
            "項目",
            "事象",
            "事象（一次）",
            "事象（二次）",
            "品番",
            "サプライヤー名",
            "不良発生連絡書発行",
            "不良発生№",
        ]
        ws.append(headers)

        # Styling header
        header_fill = PatternFill("solid", fgColor="C0C0C0")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill
            cell.border = border
            cell.font = Font(bold=True)

        wb.save(filepath)
    return filepath


def to_display_path(path: str) -> str:
    """Convert real path to display path with Japanese Yen symbol"""
    if not path:
        return ""
    # Convert backslashes to Yen symbols for display
    return path.replace("\\", "¥")


def to_real_path(path: str) -> str:
    """Convert display path with Yen symbols back to real path"""
    if not path:
        return ""
    # Convert Yen symbols back to backslashes for file operations
    return path.replace("¥", "\\")


def normalize_path(path: str) -> str:
    """Normalize path to use consistent backslashes"""
    if not path:
        return ""
    # Convert any forward slashes or mixed slashes to backslashes
    normalized = path.replace("/", "\\")
    # Replace any double backslashes with single ones
    while "\\\\" in normalized:
        normalized = normalized.replace("\\\\", "\\")
    return normalized


def is_valid_path(path: str) -> bool:
    """Check if path exists and is accessible"""
    if not path:
        return False
    try:
        real_path = to_real_path(path)
        return os.path.exists(real_path) and os.access(real_path, os.R_OK)
    except (OSError, TypeError):
        return False


def get_filename_from_path(path: str) -> str:
    """Extract filename from path (works with both ¥ and \\ separators)"""
    if not path:
        return ""
    # Normalize path first
    normalized = normalize_path(path)
    # Get filename from the end of the path
    return os.path.basename(normalized)


def format_number(value):
    """Format numeric value as integer if whole number, otherwise as string"""
    if value is None:
        return ""

    try:
        # Convert to float first to handle both int and float
        num_value = float(value)

        # Check if it's a whole number
        if num_value.is_integer():
            return str(int(num_value))
        else:
            return str(value)
    except (ValueError, TypeError):
        # If it can't be converted to number, return as string
        return str(value)


def format_excel_date(value):
    """Format Excel date value for display"""
    if value is None:
        return ""

    # Handle Excel serial dates (numbers like 421101) - but only if they're reasonable date ranges
    if isinstance(value, (int, float)):
        # Check if it's a reasonable Excel date (between 1900-01-01 and 2100-12-31)
        # Excel dates: 1 = 1900-01-01, 73050 = 2100-12-31
        if 1 <= value <= 73050:
            try:
                # Excel dates start from 1900-01-01, but Excel incorrectly treats 1900 as a leap year
                # So we need to adjust by 1 day for dates after 1900-02-28
                if value > 59:  # After 1900-02-28
                    excel_date = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=value)
                else:
                    excel_date = datetime.datetime(1899, 12, 31) + datetime.timedelta(days=value-1)

                return excel_date.strftime("%Y-%m-%d")
            except (ValueError, TypeError, OverflowError):
                return str(value)
        else:
            # If it's not a reasonable Excel date, treat as regular number
            return str(int(value)) if float(value).is_integer() else str(value)

    # Handle datetime objects
    elif isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d")

    # Handle date objects
    elif isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")

    # Handle string dates - try to parse common formats
    elif isinstance(value, str):
        value = value.strip()
        if not value:
            return ""

        # Try to parse various date formats
        date_formats = [
            "%Y-%m-%d",    # 2024-01-15
            "%Y/%m/%d",    # 2024/01/15
            "%d-%m-%Y",    # 15-01-2024
            "%d/%m/%Y",    # 15/01/2024
            "%m-%d-%Y",    # 01-15-2024
            "%m/%d/%Y",    # 01/15/2024
            "%Y年%m月%d日", # Japanese format
        ]

        for fmt in date_formats:
            try:
                parsed_date = datetime.datetime.strptime(value, fmt)
                return parsed_date.strftime("%Y-%m-%d")
            except ValueError:
                continue

        # If no format matches, return original string
        return value

    # Handle any other type
    else:
        return str(value)


def append_excel(folder, rowdata, creator):
    filepath = create_excel_if_not_exists(folder, creator)
    wb = load_workbook(filepath)
    ws = wb.active
    ws.append(rowdata)
    wb.save(filepath)
    return filepath
