import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import os, datetime, calendar, subprocess
from openpyxl import load_workbook
from excel_utils import (
    ExcelHistoryManager,
    escape_path_for_japanese_locale,
    unescape_path_for_japanese_locale,
)


# ==============================
# Japanese Labels Dictionary
# ==============================
JP_LABELS = {
    "excel_file": "Excelファイル",
    "choose_excel": "Excel検索",
    "not_selected": "(選択されていません)",
    "sheet": "シート:",
    "data_entry": "データ入力",
    "hassei_month": "発生月",
    "ruikei": "累計",
    "auto": "(自動)",
    "no": "№",
    "hassei_date": "発生日",
    "koumoku": "項目",
    "jishou": "事象",
    "ichiji": "事象（一次）",
    "niji": "事象（二次）",
    "hinban": "品番",
    "supplier": "サプライヤー名",
    "renrakusho": "不良発生連絡書発行",
    "furyo_no": "不良発生№",
    "save_add": "保存（追加）",
    "update": "更新（編集）",
    "delete": "削除",
    "clear": "クリア",
    "filter": "フィルタ",
    "excel_preview": "Excelデータプレビュー",
    "filter_result": "フィルタ結果",
    "apply_filter": "フィルタ適用",
    "close": "閉じる",
    "browse": "参照",
    "ok": "OK",
    "cancel": "キャンセル",
    "select_pdf_excel": "PDF/Excelファイル選択:",
    "select_file": "ファイル選択",
    "free_search": "フリーワード検索:",
    "from": "から:",
    "to": "まで:",
    "warning": "警告",
    "error": "エラー",
    "success": "成功",
    "deleted": "削除済み",
    "confirm": "確認",
    "confirm_delete": "本当に削除しますか？",
    "please_select_file": "ファイルを選択してください。",
    "file_not_exist": "選択したファイルが存在しません。",
    "no_data_filter": "フィルタするデータがありません。",
    "filter_applied": "フィルタ適用済み",
    "found_records": "件の一致するデータが見つかりました。",
    "file_not_found": "ファイルが見つからないか、パスが無効です。",
    "error_loading_sheets": "シート名の取得エラー:",
    "error_loading_excel": "Excelファイルの読み込みエラー:",
    "error_add_row": "行の追加エラー:",
    "error_select_row": "行の選択エラー:",
    "error_update_row": "行の更新エラー:",
    "error_delete_row": "行の削除エラー:",
    "added_ok": "データがExcelに追加されました。",
    "updated_ok": "データが更新されました。",
    "deleted_ok": "行が削除されました。",
    "pick_excel_sheet": "先にExcelファイルとシートを選択してください。",
    "pick_row_first": "先にプレビューから行を選択してください。",
    "sheet_not_found": "シート「{sheet}」が見つかりません。",
    "cannot_open_file": "ファイルを開けません:",
    "pick_excel_first": "先にExcelファイルを選択してください。",
    "date_picker_title": "日付選択",
    "weekday_mon": "月",
    "weekday_tue": "火",
    "weekday_wed": "水",
    "weekday_thu": "木",
    "weekday_fri": "金",
    "weekday_sat": "土",
    "weekday_sun": "日",
    "filter_data": "データフィルタ",
    "date_range": "発生日:",
}


# ==============================
# Simple DatePicker
# ==============================
class DatePicker(tk.Toplevel):
    def __init__(self, parent, callback, mode="ymd"):
        super().__init__(parent)
        self.callback = callback
        self.mode = mode
        self.selected = datetime.date.today()
        self.current_month = self.selected.month
        self.current_year = self.selected.year
        self.title(JP_LABELS["date_picker_title"])
        self.build()

    def build(self):
        frm = ttk.Frame(self)
        frm.pack(padx=8, pady=8)
        nav = ttk.Frame(frm)
        nav.grid(row=0, column=0, columnspan=7, pady=(0, 6))
        ttk.Button(nav, text="◀", command=self.prev_month, width=3).pack(side="left")
        self.lbl = ttk.Label(nav, text="")
        self.lbl.pack(side="left", padx=8)
        ttk.Button(nav, text="▶", command=self.next_month, width=3).pack(side="left")
        self.btns = []
        days = [
            JP_LABELS["weekday_mon"],
            JP_LABELS["weekday_tue"],
            JP_LABELS["weekday_wed"],
            JP_LABELS["weekday_thu"],
            JP_LABELS["weekday_fri"],
            JP_LABELS["weekday_sat"],
            JP_LABELS["weekday_sun"],
        ]
        for i, d in enumerate(days):
            ttk.Label(frm, text=d, width=3, anchor="center").grid(row=1, column=i)
        for r in range(6):
            rowbtn = []
            for c in range(7):
                b = tk.Button(
                    frm, width=3, command=lambda rr=r, cc=c: self.pick(rr, cc)
                )
                b.grid(row=r + 2, column=c, padx=1, pady=1)
                rowbtn.append(b)
            self.btns.append(rowbtn)
        self.update_cal()

    def update_cal(self):
        self.lbl.config(text=f"{self.current_year}-{self.current_month:02d}")
        for r in range(6):
            for c in range(7):
                self.btns[r][c].config(text="", state="disabled")
        cal = calendar.monthcalendar(self.current_year, self.current_month)
        for r, week in enumerate(cal):
            for c, day in enumerate(week):
                if day != 0:
                    self.btns[r][c].config(text=day, state="normal")

    def pick(self, r, c):
        day = int(self.btns[r][c].cget("text"))
        self.selected = datetime.date(self.current_year, self.current_month, day)
        self.callback(self.selected.strftime("%Y-%m-%d"))
        self.destroy()

    def prev_month(self):
        if self.current_month == 1:
            self.current_month = 12
            self.current_year -= 1
        else:
            self.current_month -= 1
        self.update_cal()

    def next_month(self):
        if self.current_month == 12:
            self.current_month = 1
            self.current_year += 1
        else:
            self.current_month += 1
        self.update_cal()


# ==============================
# Simplified File Selection Dialog
# ==============================
class FileSelectionDialog(tk.Toplevel):
    def __init__(self, parent, callback):
        super().__init__(parent)
        self.callback = callback
        self.result = None
        self.title(JP_LABELS["select_file"])
        self.geometry("450x150")
        self.transient(parent)
        self.grab_set()
        self.build_ui()

    def build_ui(self):
        main_frame = tk.Frame(self, padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)

        # File selection only
        tk.Label(
            main_frame, text=JP_LABELS["select_pdf_excel"], font=("Arial", 10, "bold")
        ).pack(anchor="w", pady=(0, 5))
        frame_file = tk.Frame(main_frame)
        frame_file.pack(fill="x", pady=(0, 20))

        self.entry_file = tk.Entry(frame_file, font=("Arial", 9))
        self.entry_file.pack(side="left", fill="x", expand=True)
        tk.Button(
            frame_file, text=JP_LABELS["browse"], command=self.browse_file, width=8
        ).pack(side="right", padx=(5, 0))

        # Buttons
        frame_btn = tk.Frame(main_frame)
        frame_btn.pack(pady=(10, 0))
        tk.Button(
            frame_btn,
            text=JP_LABELS["ok"],
            command=self.ok_clicked,
            width=10,
            bg="#d4edda",
        ).pack(side="left", padx=(0, 10))
        tk.Button(
            frame_btn,
            text=JP_LABELS["cancel"],
            command=self.destroy,
            width=10,
            bg="#f8d7da",
        ).pack(side="left")

        # Center the dialog
        self.center_window()

    def center_window(self):
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (self.winfo_width() // 2)
        y = (self.winfo_screenheight() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")

    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title=JP_LABELS["select_pdf_excel"],
            filetypes=[
                ("PDF files", "*.pdf"),
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*"),
            ],
        )
        if file_path:
            self.entry_file.delete(0, tk.END)
            self.entry_file.insert(0, file_path)

    def ok_clicked(self):
        file_path = self.entry_file.get().strip()

        if not file_path:
            messagebox.showwarning(
                JP_LABELS["warning"], JP_LABELS["please_select_file"]
            )
            return

        if not os.path.exists(file_path):
            messagebox.showerror(JP_LABELS["error"], JP_LABELS["file_not_exist"])
            return

        self.result = file_path
        self.callback(self.result)
        self.destroy()


# ==============================
# Filter Dialog
# ==============================
class FilterDialog(tk.Toplevel):
    def __init__(self, parent, callback, unique_data):
        super().__init__(parent)
        self.callback = callback
        self.unique_data = unique_data
        self.title(JP_LABELS["filter_data"])
        self.geometry("700x600")
        self.transient(parent)
        self.grab_set()
        self.build_ui()

    def build_ui(self):
        main_frame = tk.Frame(self, padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)

        # Create scrollable frame
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # 発生日 (Date Range)
        tk.Label(
            scrollable_frame, text=JP_LABELS["date_range"], font=("Arial", 10, "bold")
        ).grid(row=0, column=0, sticky="w", pady=(0, 5))
        date_frame = tk.Frame(scrollable_frame)
        date_frame.grid(row=0, column=1, sticky="ew", pady=(0, 10))

        tk.Label(date_frame, text=JP_LABELS["from"]).pack(side="left")
        self.entry_date_from = tk.Entry(date_frame, width=12)
        self.entry_date_from.pack(side="left", padx=2)
        tk.Button(
            date_frame,
            text="📅",
            command=lambda: self.open_date(self.entry_date_from),
            width=3,
        ).pack(side="left", padx=2)

        tk.Label(date_frame, text=JP_LABELS["to"]).pack(side="left", padx=(10, 0))
        self.entry_date_to = tk.Entry(date_frame, width=12)
        self.entry_date_to.pack(side="left", padx=2)
        tk.Button(
            date_frame,
            text="📅",
            command=lambda: self.open_date(self.entry_date_to),
            width=3,
        ).pack(side="left", padx=2)

        # 項目
        tk.Label(scrollable_frame, text=JP_LABELS["koumoku"], font=("Arial", 10)).grid(
            row=1, column=0, sticky="w", pady=5
        )
        self.cbo_koumoku = ttk.Combobox(
            scrollable_frame, values=[""] + self.unique_data["koumoku"], width=30
        )
        self.cbo_koumoku.grid(row=1, column=1, sticky="ew", pady=5)

        # 事象 (Changed to Combobox)
        tk.Label(scrollable_frame, text=JP_LABELS["jishou"], font=("Arial", 10)).grid(
            row=2, column=0, sticky="w", pady=5
        )
        self.cbo_jishou = ttk.Combobox(
            scrollable_frame, values=[""] + self.unique_data["jishou"], width=30
        )
        self.cbo_jishou.grid(row=2, column=1, sticky="ew", pady=5)

        # 事象（一次）
        tk.Label(scrollable_frame, text=JP_LABELS["ichiji"], font=("Arial", 10)).grid(
            row=3, column=0, sticky="w", pady=5
        )
        self.cbo_ichiji = ttk.Combobox(
            scrollable_frame, values=[""] + self.unique_data["ichiji"], width=30
        )
        self.cbo_ichiji.grid(row=3, column=1, sticky="ew", pady=5)

        # 事象（二次）
        tk.Label(scrollable_frame, text=JP_LABELS["niji"], font=("Arial", 10)).grid(
            row=4, column=0, sticky="w", pady=5
        )
        self.cbo_niji = ttk.Combobox(
            scrollable_frame, values=[""] + self.unique_data["niji"], width=30
        )
        self.cbo_niji.grid(row=4, column=1, sticky="ew", pady=5)

        # 品番
        tk.Label(scrollable_frame, text=JP_LABELS["hinban"], font=("Arial", 10)).grid(
            row=5, column=0, sticky="w", pady=5
        )
        self.entry_hinban = tk.Entry(scrollable_frame, width=30)
        self.entry_hinban.grid(row=5, column=1, sticky="ew", pady=5)

        # サプライヤー名
        tk.Label(scrollable_frame, text=JP_LABELS["supplier"], font=("Arial", 10)).grid(
            row=6, column=0, sticky="w", pady=5
        )
        self.cbo_supplier = ttk.Combobox(
            scrollable_frame, values=[""] + self.unique_data["suppliers"], width=30
        )
        self.cbo_supplier.grid(row=6, column=1, sticky="ew", pady=5)

        # 不良発生№
        tk.Label(scrollable_frame, text=JP_LABELS["furyo_no"], font=("Arial", 10)).grid(
            row=7, column=0, sticky="w", pady=5
        )
        self.entry_furyo_no = tk.Entry(scrollable_frame, width=30)
        self.entry_furyo_no.grid(row=7, column=1, sticky="ew", pady=5)

        # Separator
        ttk.Separator(scrollable_frame, orient="horizontal").grid(
            row=8, column=0, columnspan=2, sticky="ew", pady=10
        )

        # Free Search
        tk.Label(
            scrollable_frame, text=JP_LABELS["free_search"], font=("Arial", 10, "bold")
        ).grid(row=9, column=0, sticky="w", pady=5)
        self.entry_free_search = tk.Entry(scrollable_frame, width=30)
        self.entry_free_search.grid(row=9, column=1, sticky="ew", pady=5)

        scrollable_frame.grid_columnconfigure(1, weight=1)

        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Buttons
        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=10)
        tk.Button(
            btn_frame,
            text=JP_LABELS["apply_filter"],
            command=self.apply_filter,
            width=12,
            bg="#d4edda",
        ).pack(side="left", padx=5)
        tk.Button(
            btn_frame,
            text=JP_LABELS["clear"],
            command=self.clear_filter,
            width=12,
            bg="#fff3cd",
        ).pack(side="left", padx=5)
        tk.Button(
            btn_frame,
            text=JP_LABELS["close"],
            command=self.destroy,
            width=12,
            bg="#f8d7da",
        ).pack(side="left", padx=5)

        self.center_window()

    def center_window(self):
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (self.winfo_width() // 2)
        y = (self.winfo_screenheight() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")

    def open_date(self, entry):
        DatePicker(self, self.set_date(entry), mode="ymd")

    def set_date(self, entry):
        def cb(val):
            entry.delete(0, tk.END)
            entry.insert(0, val)

        return cb

    def clear_filter(self):
        self.entry_date_from.delete(0, tk.END)
        self.entry_date_to.delete(0, tk.END)
        self.cbo_koumoku.set("")
        self.cbo_jishou.set("")  # Changed from entry to combobox
        self.cbo_ichiji.set("")
        self.cbo_niji.set("")
        self.entry_hinban.delete(0, tk.END)
        self.cbo_supplier.set("")
        self.entry_furyo_no.delete(0, tk.END)
        self.entry_free_search.delete(0, tk.END)

    def apply_filter(self):
        filters = {
            "date_from": self.entry_date_from.get().strip(),
            "date_to": self.entry_date_to.get().strip(),
            "koumoku": self.cbo_koumoku.get().strip(),
            "jishou": self.cbo_jishou.get().strip(),  # Changed from entry to combobox
            "ichiji": self.cbo_ichiji.get().strip(),
            "niji": self.cbo_niji.get().strip(),
            "hinban": self.entry_hinban.get().strip(),
            "supplier": self.cbo_supplier.get().strip(),
            "furyo_no": self.entry_furyo_no.get().strip(),
            "free_search": self.entry_free_search.get().strip(),
        }
        self.callback(filters)
        self.destroy()


# ==============================
# For record history of file
# ==============================
class ExcelHistoryDialog(tk.Toplevel):
    def __init__(self, parent, callback, history_manager):
        super().__init__(parent)
        self.callback = callback
        self.history_manager = history_manager
        self.result = None
        self.title("Excel履歴")  # Excel History
        self.geometry("600x400")
        self.transient(parent)
        self.grab_set()
        self.build_ui()

    def build_ui(self):
        main_frame = tk.Frame(self, padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)

        # Title
        tk.Label(
            main_frame, text="最近使用したExcelファイル", font=("Arial", 12, "bold")
        ).pack(anchor="w", pady=(0, 10))

        # History list with scrollbar
        list_frame = tk.Frame(main_frame)
        list_frame.pack(fill="both", expand=True, pady=(0, 10))

        # Listbox with scrollbar
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side="right", fill="y")

        self.listbox = tk.Listbox(
            list_frame, yscrollcommand=scrollbar.set, font=("Arial", 9)
        )
        self.listbox.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.listbox.yview)

        # Double click to select
        self.listbox.bind("<Double-Button-1>", self.on_double_click)

        # Load history items
        self.load_history()

        # Buttons frame
        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(fill="x", pady=(10, 0))

        tk.Button(
            btn_frame, text="開く", command=self.open_selected, width=10, bg="#d4edda"
        ).pack(side="left", padx=(0, 5))
        tk.Button(
            btn_frame, text="削除", command=self.remove_selected, width=10, bg="#f8d7da"
        ).pack(side="left", padx=5)
        tk.Button(
            btn_frame, text="全削除", command=self.clear_all, width=10, bg="#f8d7da"
        ).pack(side="left", padx=5)
        tk.Button(
            btn_frame, text="キャンセル", command=self.destroy, width=10, bg="#f8f9fa"
        ).pack(side="right")

        self.center_window()

    def center_window(self):
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (self.winfo_width() // 2)
        y = (self.winfo_screenheight() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")

    def load_history(self):
        """Load history items into listbox"""
        self.listbox.delete(0, tk.END)
        items = self.history_manager.items()

        if not items:
            self.listbox.insert(tk.END, "（履歴がありません）")
            return

        for i, path in enumerate(items):
            # Show filename and path
            filename = os.path.basename(path)
            display_text = f"{i+1}. {filename}"

            # Check if file exists
            if not os.path.exists(path):
                display_text += " (ファイルが見つかりません)"

            self.listbox.insert(tk.END, display_text)

            # Store full path as data
            self.listbox.insert(tk.END, f"   📁 {path}")

    def on_double_click(self, event):
        """Handle double click on listbox item"""
        self.open_selected()

    def open_selected(self):
        """Open selected file"""
        selection = self.listbox.curselection()
        if not selection:
            messagebox.showwarning("警告", "ファイルを選択してください。")
            return

        # Get the selected index (considering display format)
        selected_idx = selection[0]
        items = self.history_manager.items()

        if not items:
            return

        # Calculate actual file index (every file takes 2 lines in listbox)
        file_idx = selected_idx // 2

        if file_idx >= len(items):
            return

        file_path = items[file_idx]

        # Check if file exists
        if not os.path.exists(file_path):
            response = messagebox.askyesno(
                "ファイルが見つかりません",
                f"ファイルが見つかりません：\n{file_path}\n\n履歴から削除しますか？",
            )
            if response:
                self.history_manager.remove(file_path)
                self.load_history()
            return

        # Call callback with selected file
        self.result = file_path
        self.callback(file_path)
        self.destroy()

    def remove_selected(self):
        """Remove selected file from history"""
        selection = self.listbox.curselection()
        if not selection:
            messagebox.showwarning("警告", "削除するファイルを選択してください。")
            return

        selected_idx = selection[0]
        items = self.history_manager.items()

        if not items:
            return

        file_idx = selected_idx // 2

        if file_idx >= len(items):
            return

        file_path = items[file_idx]
        filename = os.path.basename(file_path)

        if messagebox.askyesno("確認", f"履歴から削除しますか？\n{filename}"):
            self.history_manager.remove(file_path)
            self.load_history()

    def clear_all(self):
        """Clear all history"""
        if messagebox.askyesno("確認", "すべての履歴を削除しますか？"):
            self.history_manager.clear()
            self.load_history()


# ==============================
# Tab2Entry (Master-Detail Layout with Filter)
# ==============================
class Tab2Entry:
    def __init__(self, parent, app):
        self.app = app
        self.excel_path = None
        self.selected_sheet = None
        self.selected_row = None  # Excel row index
        self.history_manager = ExcelHistoryManager()  # Initialize history manager
        self.all_data = []  # Store all Excel data for filtering
        self.header_row = None  # Dynamic header row detection
        self.data_start_row = None  # Dynamic data start row
        self.is_filter_mode = False  # Track if we're in filter mode

        # Root containers: left (form) and right (preview)
        root = tk.Frame(parent)
        root.pack(fill="both", expand=True)
        self.root = root

        left = tk.Frame(root, padx=10, pady=10)
        right = tk.Frame(root, padx=10, pady=10)
        left.pack(side="left", fill="y")
        right.pack(side="right", fill="both", expand=True)

        # ------------------------------
        # Left: File pick + Sheet selector + Form + Actions
        # ------------------------------
        sec_file = tk.LabelFrame(left, text=JP_LABELS["excel_file"], padx=8, pady=8)
        sec_file.pack(fill="x")

        # File selection row
        file_row = tk.Frame(sec_file)
        file_row.pack(fill="x", pady=(0, 5))
        self.btn_choose_file = tk.Button(
            file_row, text=JP_LABELS["choose_excel"], command=self.choose_file
        )
        self.btn_choose_file.pack(side="left")
        self.lbl_file = tk.Label(
            file_row, text=JP_LABELS["not_selected"], width=25, anchor="w"
        )
        self.lbl_file.pack(side="left", padx=8)
        self.btn_history = tk.Button(
            file_row, text="履歴📋", command=self.show_history, width=3, bg="#e9ecef"
        )
        self.btn_history.pack(side="left", padx=2)

        # Sheet selection row
        sheet_row = tk.Frame(sec_file)
        sheet_row.pack(fill="x")
        tk.Label(sheet_row, text=JP_LABELS["sheet"], width=8, anchor="w").pack(
            side="left"
        )
        self.cbo_sheet = ttk.Combobox(sheet_row, width=25, state="readonly")
        self.cbo_sheet.pack(side="left", padx=(5, 0))
        self.cbo_sheet.bind("<<ComboboxSelected>>", self.on_sheet_change)

        sec_form = tk.LabelFrame(left, text=JP_LABELS["data_entry"], padx=8, pady=8)
        sec_form.pack(fill="x", pady=(10, 8))

        # Defaults (combobox values)
        self.default_koumoku = [
            "加工(未加工/誤加工)不良",
            "塗装・鍍金不良",
            "塗装不良",
            "外観・形状不良",
            "外観不良",
            "寸法不良",
            "架台塗装剥がれ",
            "機能・性能・動作不良",
            "機能不良",
            "現品相違",
            "組立不良",
            "表示間違い",
        ]
        self.default_ichiji = [
            "エアー漏れ",
            "ネジ・タップ不良",
            "仕上げ不良",
            "仕様間違い",
            "外観不良",
            "寸法不良",
            "機能・性能・動作不良",
            "組立不良",
        ]
        self.default_niji = [
            "キズ",
            "サビ不良",
            "その他",
            "ネジタップ不良",
            "ネジ不良",
            "はんだ不良",
            "メッキ不良",
            "ワレ",
            "加工不良",
            "基盤不良",
            "変形",
            "寸法不良",
            "打コン",
            "未加工",
            "溶接不良",
            "異物",
            "誤加工",
            "鋳造不良",
        ]
        self.default_suppliers = [
            "WAKO",
            "キョウセイ",
            "タナカマシ―ナリー",
            "ネイティング",
            "ファインテクノ・タケダ",
            "プランネットエンジニアリング",
            "ミヤケ工業",
            "ヤハタ",
            "ヤマニゴム",
            "三光電業",
            "三谷製作所",
            "中国ゴム工業",
            "五敬工業",
            "人見特装",
            "佐藤商事",
            "倉敷レーザー",
            "光南工業所",
            "八光",
            "北陽商事",
            "吉沢製作所",
            "坪井鉄工所",
            "姫路鍍金工業所",
            "岡鉄工所",
            "平井工作所",
            "平和アルキャス",
            "御幸鉄工所",
            "新生産業",
            "日本ケミカル",
            "旭金属工業",
            "東京工販",
            "椿本興業",
            "池田精工",
            "真鉄工",
            "石原パッキング工業",
            "西大寺塗装工業所",
            "連島工業",
            "野口工業",
            "除振開発",
            "頼鉄工",
        ]

        row = 0

        def add_row(label, widget):
            nonlocal row
            tk.Label(sec_form, text=label, width=12, anchor="w").grid(
                row=row, column=0, sticky="w", pady=2
            )
            widget.grid(row=row, column=1, sticky="ew", pady=2)
            row += 1

        sec_form.grid_columnconfigure(1, weight=1)

        # 発生月
        self.entry_hassei_month = tk.Entry(sec_form, width=20)
        self.btn_cal1 = tk.Button(
            sec_form, text="📅", command=lambda: self.open_date(self.entry_hassei_month)
        )
        tk.Label(sec_form, text=JP_LABELS["hassei_month"], width=12, anchor="w").grid(
            row=row, column=0, sticky="w", pady=2
        )
        self.entry_hassei_month.grid(row=row, column=1, sticky="ew", pady=2)
        self.btn_cal1.grid(row=row, column=2, padx=2)
        row += 1

        # 累計 (auto)
        self.lbl_ruikei = tk.Label(sec_form, text=JP_LABELS["auto"], anchor="w")
        add_row(JP_LABELS["ruikei"], self.lbl_ruikei)

        # № (int)
        vcmd = (sec_form.register(self.validate_int), "%P")
        self.entry_no = tk.Entry(sec_form, validate="key", validatecommand=vcmd)
        add_row(JP_LABELS["no"], self.entry_no)

        # 発生日
        self.entry_date = tk.Entry(sec_form, width=20)
        self.btn_cal2 = tk.Button(
            sec_form, text="📅", command=lambda: self.open_date(self.entry_date)
        )
        tk.Label(sec_form, text=JP_LABELS["hassei_date"], width=12, anchor="w").grid(
            row=row, column=0, sticky="w", pady=2
        )
        self.entry_date.grid(row=row, column=1, sticky="ew", pady=2)
        self.btn_cal2.grid(row=row, column=2, padx=2)
        row += 1

        # 項目
        self.cbo_koumoku = ttk.Combobox(sec_form, values=self.default_koumoku, width=30)
        add_row(JP_LABELS["koumoku"], self.cbo_koumoku)

        # 事象
        self.entry_jishou = tk.Entry(sec_form, width=30)
        add_row(JP_LABELS["jishou"], self.entry_jishou)

        # 事象（一次）
        self.cbo_ichiji = ttk.Combobox(sec_form, values=self.default_ichiji, width=30)
        add_row(JP_LABELS["ichiji"], self.cbo_ichiji)

        # 事象（二次）
        self.cbo_niji = ttk.Combobox(sec_form, values=self.default_niji, width=30)
        add_row(JP_LABELS["niji"], self.cbo_niji)

        # 品番
        self.entry_hinban = tk.Entry(sec_form, width=30)
        add_row(JP_LABELS["hinban"], self.entry_hinban)

        # サプライヤー名
        self.cbo_supplier = ttk.Combobox(
            sec_form, values=self.default_suppliers, width=30
        )
        add_row(JP_LABELS["supplier"], self.cbo_supplier)

        # 不良発生連絡書発行 (NEW FIELD)
        frame_renrakusho = tk.Frame(sec_form)
        frame_renrakusho.grid(row=row, column=1, sticky="ew", pady=2)
        frame_renrakusho.grid_columnconfigure(0, weight=1)

        self.entry_renrakusho = tk.Entry(frame_renrakusho)
        self.entry_renrakusho.grid(row=0, column=0, sticky="ew")

        self.btn_browse = tk.Button(
            frame_renrakusho, text="📁", command=self.browse_renrakusho, width=3
        )
        self.btn_browse.grid(row=0, column=1, padx=2)

        self.btn_open_file = tk.Button(
            frame_renrakusho, text="📄", command=self.open_renrakusho_file, width=3
        )
        self.btn_open_file.grid(row=0, column=2, padx=2)
        self.btn_open_file.config(state="disabled")  # disabled sampai ada file

        tk.Label(sec_form, text=JP_LABELS["renrakusho"], width=12, anchor="w").grid(
            row=row, column=0, sticky="w", pady=2
        )
        row += 1

        # 不良発生№
        self.entry_furyo_no = tk.Entry(sec_form, width=20)
        add_row(JP_LABELS["furyo_no"], self.entry_furyo_no)

        # Action buttons
        sec_actions = tk.Frame(left)
        sec_actions.pack(fill="x", pady=(6, 0))
        self.btn_add = tk.Button(
            sec_actions, text=JP_LABELS["save_add"], command=self.add_row, bg="#d4edda"
        )
        self.btn_add.pack(side="left")
        self.btn_update = tk.Button(
            sec_actions, text=JP_LABELS["update"], command=self.update_row, bg="#fff3cd"
        )
        self.btn_update.pack(side="left", padx=6)
        self.btn_delete = tk.Button(
            sec_actions, text=JP_LABELS["delete"], command=self.delete_row, bg="#f8d7da"
        )
        self.btn_delete.pack(side="left")
        self.btn_clear = tk.Button(
            sec_actions, text=JP_LABELS["clear"], command=self.clear_form
        )
        self.btn_clear.pack(side="left", padx=6)
        self.btn_filter = tk.Button(
            sec_actions,
            text=JP_LABELS["filter"],
            command=self.open_filter_dialog,
            bg="#e2e3e5",
        )
        self.btn_filter.pack(side="left", padx=6)

        # Store UI elements for enable/disable
        self.ui_elements = {
            "file_section": [self.btn_choose_file, self.cbo_sheet],
            "form_section": [
                self.entry_hassei_month,
                self.btn_cal1,
                self.entry_no,
                self.entry_date,
                self.btn_cal2,
                self.cbo_koumoku,
                self.entry_jishou,
                self.cbo_ichiji,
                self.cbo_niji,
                self.entry_hinban,
                self.cbo_supplier,
                self.btn_browse,
                self.entry_furyo_no,
            ],
            "action_buttons": [
                self.btn_add,
                self.btn_update,
                self.btn_delete,
                self.btn_clear,
            ],
            "always_enabled": [
                self.btn_filter,
                self.btn_open_file,
                self.entry_renrakusho,
            ],  # These stay enabled in filter mode
        }

        # ------------------------------
        # Right: Tabbed Preview (Excel Data + Filter Result)
        # ------------------------------
        self.notebook = ttk.Notebook(right)
        self.notebook.pack(fill="both", expand=True)

        # Tab 1: Excel Data Preview
        tab1 = ttk.Frame(self.notebook)
        self.notebook.add(tab1, text=JP_LABELS["excel_preview"])

        sec_preview1 = tk.LabelFrame(
            tab1, text=JP_LABELS["excel_preview"], padx=8, pady=8
        )
        sec_preview1.pack(fill="both", expand=True)

        container1 = tk.Frame(sec_preview1)
        container1.pack(fill="both", expand=True)

        self.tree = ttk.Treeview(container1, show="headings")
        vsb1 = ttk.Scrollbar(container1, orient="vertical", command=self.tree.yview)
        hsb1 = ttk.Scrollbar(container1, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb1.set, xscrollcommand=hsb1.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb1.grid(row=0, column=1, sticky="ns")
        hsb1.grid(row=1, column=0, sticky="ew")

        container1.grid_rowconfigure(0, weight=1)
        container1.grid_columnconfigure(0, weight=1)

        # Tab 2: Filter Result
        tab2 = ttk.Frame(self.notebook)
        self.notebook.add(tab2, text=JP_LABELS["filter_result"])

        sec_preview2 = tk.LabelFrame(
            tab2, text=JP_LABELS["filter_result"], padx=8, pady=8
        )
        sec_preview2.pack(fill="both", expand=True)

        container2 = tk.Frame(sec_preview2)
        container2.pack(fill="both", expand=True)

        self.filter_tree = ttk.Treeview(container2, show="headings")
        vsb2 = ttk.Scrollbar(
            container2, orient="vertical", command=self.filter_tree.yview
        )
        hsb2 = ttk.Scrollbar(
            container2, orient="horizontal", command=self.filter_tree.xview
        )
        self.filter_tree.configure(yscrollcommand=vsb2.set, xscrollcommand=hsb2.set)

        self.filter_tree.grid(row=0, column=0, sticky="nsew")
        vsb2.grid(row=0, column=1, sticky="ns")
        hsb2.grid(row=1, column=0, sticky="ew")

        container2.grid_rowconfigure(0, weight=1)
        container2.grid_columnconfigure(0, weight=1)

        # Bind select events
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        self.filter_tree.bind("<<TreeviewSelect>>", self.on_filter_tree_select)

        # Bind tab change event
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_change)

    # ==============================
    # UI State Management
    # ==============================
    def set_ui_state(self, enabled=True):
        """Enable/disable UI elements based on filter mode"""
        state = "normal" if enabled else "disabled"

        # File section and form section
        for element in (
            self.ui_elements["file_section"] + self.ui_elements["form_section"]
        ):
            if hasattr(element, "config"):
                element.config(state=state)

        # Action buttons (except filter and open file)
        for element in self.ui_elements["action_buttons"]:
            if hasattr(element, "config"):
                element.config(state=state)

        # Always enabled elements stay enabled
        for element in self.ui_elements["always_enabled"]:
            if hasattr(element, "config"):
                element.config(state="normal")

    def update_button_states(self):
        """Update button states based on selection"""
        if self.selected_row is not None:
            # Data is selected - disable add button, enable update/delete
            self.btn_add.config(state="disabled")
            self.btn_update.config(state="normal")
            self.btn_delete.config(state="normal")
        else:
            # No data selected - enable add button, disable update/delete
            self.btn_add.config(state="normal")
            self.btn_update.config(state="disabled")
            self.btn_delete.config(state="disabled")

    def on_tab_change(self, event):
        """Handle tab change events"""
        selected_tab = event.widget.tab("current")["text"]

        if selected_tab == JP_LABELS["filter_result"]:
            self.is_filter_mode = True
            self.set_ui_state(False)  # Disable UI elements
        else:
            self.is_filter_mode = False
            self.set_ui_state(True)  # Enable UI elements
            self.update_button_states()  # Update button states based on selection

    # ==============================
    # Dynamic Table Detection
    # ==============================
    def find_table_position(self, ws):
        """Dynamically find header row and data start position"""
        # Search for header indicators in first 20 rows
        header_indicators = ["発生月", "累計", "№", "発生日", "項目", "事象"]

        for row_num in range(1, 21):  # Check first 20 rows
            row_values = [cell.value for cell in ws[row_num]]
            row_text = [str(val).strip() if val else "" for val in row_values]

            # Check if this row contains header indicators
            matches = sum(
                1
                for indicator in header_indicators
                if any(indicator in text for text in row_text)
            )

            if matches >= 3:  # If at least 3 indicators found
                self.header_row = row_num
                self.data_start_row = row_num + 1
                return True

        # Fallback to default positions if not found
        self.header_row = 3
        self.data_start_row = 4
        return False

    # ==============================
    # Sheet Management
    # ==============================
    def on_sheet_change(self, event=None):
        """Handle sheet selection change"""
        if self.excel_path and self.cbo_sheet.get():
            self.selected_sheet = self.cbo_sheet.get()
            self.load_excel_to_tree()

    def load_sheet_names(self):
        """Load available sheet names into combobox"""
        if not self.excel_path:
            return

        try:
            wb = load_workbook(self.excel_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            self.cbo_sheet["values"] = sheet_names
            if sheet_names:
                self.cbo_sheet.set(sheet_names[0])  # Select first sheet by default
                self.selected_sheet = sheet_names[0]

        except Exception as e:
            messagebox.showerror(
                JP_LABELS["error"], f"{JP_LABELS['error_loading_sheets']} {str(e)}"
            )

    # ==============================
    # Extract Unique Data for Filter
    # ==============================
    def extract_unique_data(self):
        """Extract unique values from Excel data for filter comboboxes"""
        if not self.all_data:
            return {
                "koumoku": [],
                "jishou": [],
                "ichiji": [],
                "niji": [],
                "suppliers": [],
            }

        # Column indices (adjust based on your Excel structure)
        koumoku_col = 4  # 項目
        jishou_col = 5  # 事象
        ichiji_col = 6  # 事象（一次）
        niji_col = 7  # 事象（二次）
        supplier_col = 9  # サプライヤー名

        unique_data = {
            "koumoku": set(),
            "jishou": set(),
            "ichiji": set(),
            "niji": set(),
            "suppliers": set(),
        }

        for row in self.all_data:
            # Extract unique values, skip None and empty strings
            if (
                len(row) > koumoku_col
                and row[koumoku_col]
                and str(row[koumoku_col]).strip()
            ):
                unique_data["koumoku"].add(str(row[koumoku_col]).strip())

            if (
                len(row) > jishou_col
                and row[jishou_col]
                and str(row[jishou_col]).strip()
            ):
                unique_data["jishou"].add(str(row[jishou_col]).strip())

            if (
                len(row) > ichiji_col
                and row[ichiji_col]
                and str(row[ichiji_col]).strip()
            ):
                unique_data["ichiji"].add(str(row[ichiji_col]).strip())

            if len(row) > niji_col and row[niji_col] and str(row[niji_col]).strip():
                unique_data["niji"].add(str(row[niji_col]).strip())

            if (
                len(row) > supplier_col
                and row[supplier_col]
                and str(row[supplier_col]).strip()
            ):
                unique_data["suppliers"].add(str(row[supplier_col]).strip())

        # Convert sets to sorted lists
        return {
            "koumoku": sorted(list(unique_data["koumoku"])),
            "jishou": sorted(list(unique_data["jishou"])),
            "ichiji": sorted(list(unique_data["ichiji"])),
            "niji": sorted(list(unique_data["niji"])),
            "suppliers": sorted(list(unique_data["suppliers"])),
        }

    # ==============================
    # Filter Functions
    # ==============================
    def open_filter_dialog(self):
        if not self.excel_path:
            messagebox.showwarning(JP_LABELS["warning"], JP_LABELS["pick_excel_first"])
            return

        # Extract unique data from current Excel sheet
        unique_data = self.extract_unique_data()
        FilterDialog(self.root, self.apply_filter, unique_data)

    def apply_filter(self, filters):
        if not self.all_data:
            messagebox.showwarning(JP_LABELS["warning"], JP_LABELS["no_data_filter"])
            return

        filtered_data = []

        for row_data in self.all_data:
            match = True

            # Convert row_data to strings for comparison, handle None values
            row_str = [str(cell) if cell is not None else "" for cell in row_data]

            # Date range filter (assuming date is in column 3 - 発生日)
            if filters["date_from"] or filters["date_to"]:
                date_val = row_str[3] if len(row_str) > 3 else ""
                if filters["date_from"] and date_val < filters["date_from"]:
                    match = False
                if filters["date_to"] and date_val > filters["date_to"]:
                    match = False

            # Specific field filters
            if (
                filters["koumoku"]
                and filters["koumoku"].lower() not in row_str[4].lower()
                if len(row_str) > 4
                else False
            ):
                match = False
            if (
                filters["jishou"]
                and filters["jishou"].lower() not in row_str[5].lower()
                if len(row_str) > 5
                else False
            ):
                match = False
            if (
                filters["ichiji"]
                and filters["ichiji"].lower() not in row_str[6].lower()
                if len(row_str) > 6
                else False
            ):
                match = False
            if (
                filters["niji"] and filters["niji"].lower() not in row_str[7].lower()
                if len(row_str) > 7
                else False
            ):
                match = False
            if (
                filters["hinban"]
                and filters["hinban"].lower() not in row_str[8].lower()
                if len(row_str) > 8
                else False
            ):
                match = False
            if (
                filters["supplier"]
                and filters["supplier"].lower() not in row_str[9].lower()
                if len(row_str) > 9
                else False
            ):
                match = False
            if (
                filters["furyo_no"]
                and filters["furyo_no"].lower() not in row_str[11].lower()
                if len(row_str) > 11
                else False
            ):
                match = False

            # Free search (search in all columns)
            if filters["free_search"]:
                free_match = False
                search_term = filters["free_search"].lower()
                for cell in row_str:
                    if search_term in cell.lower():
                        free_match = True
                        break
                if not free_match:
                    match = False

            if match:
                filtered_data.append(row_data)

        # Display filtered results in filter_tree
        self.display_filtered_data(filtered_data)

        # Switch to filter result tab
        self.notebook.select(1)

        messagebox.showinfo(
            JP_LABELS["filter_applied"],
            f"{len(filtered_data)}{JP_LABELS['found_records']}",
        )

    def display_filtered_data(self, filtered_data):
        # Clear filter tree
        self.filter_tree.delete(*self.filter_tree.get_children())

        if not filtered_data:
            return

        # Use same headers as main tree
        headers = self.tree["columns"]
        self.filter_tree["columns"] = headers

        # Set column properties
        for i, h in enumerate(headers):
            col_values = []
            for row in filtered_data:
                if i < len(row) and row[i] is not None:
                    col_values.append(str(row[i]))
                else:
                    col_values.append("")

            longest = max([len(str(h))] + [len(v) for v in col_values] + [10])
            width = min(max(100, longest * 9), 600)

            self.filter_tree.heading(h, text=h)
            self.filter_tree.column(h, width=width, anchor="w", stretch=False)

        # Insert filtered data
        for row in filtered_data:
            # Convert None to empty string
            clean_row = [cell if cell is not None else "" for cell in row]
            self.filter_tree.insert("", tk.END, values=clean_row)

    def on_filter_tree_select(self, event):
        # Handle selection in filter tree - fill form but don't change selected_row
        sel = self.filter_tree.selection()
        if not sel:
            return

        try:
            vals = self.filter_tree.item(sel[0], "values")
            # Fill form with selected data but don't set selected_row (read-only mode)
            self.fill_form_with_data(vals, read_only=True)

        except Exception as e:
            messagebox.showerror(
                JP_LABELS["error"], f"{JP_LABELS['error_select_row']} {str(e)}"
            )

    # ==============================
    # File Selection Functions
    # ==============================
    def browse_renrakusho(self):
        """Open file selection dialog"""

        def callback(file_path):
            self.entry_renrakusho.delete(0, tk.END)
            self.entry_renrakusho.insert(0, file_path)
            # Enable open button if file exists
            if file_path and os.path.exists(file_path):
                self.btn_open_file.config(state="normal")
            else:
                self.btn_open_file.config(state="disabled")

        FileSelectionDialog(self.root, callback)

    def open_renrakusho_file(self):
        """Open the selected file"""
        escaped_path = self.entry_renrakusho.get().strip()
        if not escaped_path:
            messagebox.showwarning(JP_LABELS["warning"], JP_LABELS["file_not_found"])
            return

        # Unescape path sebelum dibuka
        file_path = unescape_path_for_japanese_locale(escaped_path)

        if not os.path.exists(file_path):
            messagebox.showwarning(JP_LABELS["warning"], JP_LABELS["file_not_found"])
            return

        try:
            if os.name == "nt":  # Windows
                os.startfile(file_path)
            elif os.name == "posix":  # macOS/Linux
                subprocess.call(
                    [
                        (
                            "open"
                            if "darwin" in os.uname().sysname.lower()
                            else "xdg-open"
                        ),
                        file_path,
                    ]
                )
        except Exception as e:
            messagebox.showerror(
                JP_LABELS["error"], f"{JP_LABELS['cannot_open_file']} {str(e)}"
            )

    # ==============================
    # File ops + tree rendering
    # ==============================
    def choose_file(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx")], title=JP_LABELS["choose_excel"]
        )
        if path:
            self.excel_path = path
            self.lbl_file.config(text=os.path.basename(path))

            # Add to history
            self.history_manager.add(path)

            self.load_sheet_names()  # Load available sheets
            self.load_excel_to_tree()

    def show_history(self):
        """Show Excel file history dialog"""

        def on_file_selected(file_path):
            self.excel_path = file_path
            self.lbl_file.config(text=os.path.basename(file_path))
            self.load_sheet_names()
            self.load_excel_to_tree()

        ExcelHistoryDialog(self.root, on_file_selected, self.history_manager)

    def load_excel_to_tree(self):
        if not self.excel_path or not os.path.exists(self.excel_path):
            return

        if not self.selected_sheet:
            return

        if self.excel_path and os.path.exists(self.excel_path):
            # Add to history when successfully opened
            self.history_manager.add(self.excel_path)

        try:
            wb = load_workbook(self.excel_path)

            # Select the specified sheet
            if self.selected_sheet in wb.sheetnames:
                ws = wb[self.selected_sheet]
            else:
                messagebox.showerror(
                    JP_LABELS["error"],
                    JP_LABELS["sheet_not_found"].format(sheet=self.selected_sheet),
                )
                return

            # Dynamically find table position
            self.find_table_position(ws)

            # Get headers from detected header row
            headers = [cell.value for cell in ws[self.header_row]]
            # Clean headers (remove None values)
            headers = [
                str(h) if h is not None else f"Column_{i}"
                for i, h in enumerate(headers)
            ]

            self.tree["columns"] = headers
            self.tree.delete(*self.tree.get_children())

            # Store all data for filtering
            self.all_data = []

            # Auto-adjust column widths based on content
            for i, h in enumerate(headers):
                col_values = []
                for row in ws.iter_rows(min_row=self.data_start_row, values_only=True):
                    if i < len(row) and row[i] is not None:
                        col_values.append(str(row[i]))
                    else:
                        col_values.append("")

                # Calculate optimal width
                longest = max([len(str(h))] + [len(v) for v in col_values] + [10])
                width = min(max(100, longest * 9), 600)

                self.tree.heading(h, text=h)
                self.tree.column(h, width=width, anchor="w", stretch=False)

            # Load data starting from detected data row
            for row in ws.iter_rows(min_row=self.data_start_row, values_only=True):
                # Convert None to empty string for display
                clean_row = [cell if cell is not None else "" for cell in row]
                self.tree.insert("", tk.END, values=clean_row)
                # Store original data (with None values) for filtering
                self.all_data.append(row)

            # Update ruikei label (jumlah data)
            data_count = len(self.all_data)
            self.lbl_ruikei.config(text=str(data_count))
            self.selected_row = None
            self.update_button_states()  # Update button states

        except Exception as e:
            messagebox.showerror(
                JP_LABELS["error"], f"{JP_LABELS['error_loading_excel']} {str(e)}"
            )

    # ==============================
    # Date Picker helpers
    # ==============================
    def open_date(self, entry):
        DatePicker(self.root, self.set_date(entry), mode="ymd")

    def set_date(self, entry):
        def cb(val):
            entry.delete(0, tk.END)
            entry.insert(0, val)

        return cb

    # ==============================
    # Form helpers / validation
    # ==============================
    def clear_form(self):
        self.entry_hassei_month.delete(0, tk.END)
        self.lbl_ruikei.config(text=JP_LABELS["auto"])
        self.entry_no.delete(0, tk.END)
        self.entry_date.delete(0, tk.END)
        self.cbo_koumoku.set("")  # Changed from default value to empty string
        self.entry_jishou.delete(0, tk.END)
        self.cbo_ichiji.set("")
        self.cbo_niji.set("")
        self.entry_hinban.delete(0, tk.END)
        self.cbo_supplier.set("")
        self.entry_renrakusho.delete(0, tk.END)
        self.btn_open_file.config(state="disabled")
        self.entry_furyo_no.delete(0, tk.END)
        self.selected_row = None
        self.update_button_states()  # Update button states after clearing

    def validate_int(self, value):
        if value == "":
            return True
        return value.isdigit()

    def fill_form_with_data(self, vals, read_only=False):
        """Fill form with data from tree selection"""
        self.entry_hassei_month.delete(0, tk.END)
        self.entry_hassei_month.insert(0, vals[0] if len(vals) > 0 else "")

        self.entry_no.delete(0, tk.END)
        self.entry_no.insert(0, vals[2] if len(vals) > 2 else "")

        self.entry_date.delete(0, tk.END)
        self.entry_date.insert(0, vals[3] if len(vals) > 3 else "")

        self.cbo_koumoku.set(vals[4] if len(vals) > 4 else "")

        self.entry_jishou.delete(0, tk.END)
        self.entry_jishou.insert(0, vals[5] if len(vals) > 5 else "")

        self.cbo_ichiji.set(vals[6] if len(vals) > 6 else "")
        self.cbo_niji.set(vals[7] if len(vals) > 7 else "")

        self.entry_hinban.delete(0, tk.END)
        self.entry_hinban.insert(0, vals[8] if len(vals) > 8 else "")

        self.cbo_supplier.set(vals[9] if len(vals) > 9 else "")

        # 不良発生連絡書発行
        self.entry_renrakusho.delete(0, tk.END)
        renrakusho_path = vals[10] if len(vals) > 10 else ""
        self.entry_renrakusho.insert(0, renrakusho_path)

        # Enable/disable open button based on file existence
        if renrakusho_path and os.path.exists(renrakusho_path):
            self.btn_open_file.config(state="normal")
        else:
            self.btn_open_file.config(state="disabled")

        self.entry_furyo_no.delete(0, tk.END)
        self.entry_furyo_no.insert(0, vals[11] if len(vals) > 11 else "")

    # ==============================
    # CRUD ops on Excel
    # ==============================
    def add_row(self):
        if not self.excel_path or not self.selected_sheet:
            messagebox.showwarning(JP_LABELS["warning"], JP_LABELS["pick_excel_sheet"])
            return

        try:
            wb = load_workbook(self.excel_path)
            ws = wb[self.selected_sheet]

            # ruikei = jumlah data existing + 1
            current_data_count = len(self.all_data)
            ruikei = current_data_count + 1

            # Escape path untuk kompatibilitas Jepang
            escaped_renrakusho_path = escape_path_for_japanese_locale(
                self.entry_renrakusho.get() or ""
            )

            vals = [
                self.entry_hassei_month.get() or "",  # Convert empty to ""
                ruikei,  # 累計 (col 2)
                self.entry_no.get() or "",  # № (col 3)
                self.entry_date.get() or "",
                self.cbo_koumoku.get() or "",
                self.entry_jishou.get() or "",
                self.cbo_ichiji.get() or "",
                self.cbo_niji.get() or "",
                self.entry_hinban.get() or "",
                self.cbo_supplier.get() or "",
                escaped_renrakusho_path,  # 不良発生連絡書発行 (col 11) - dalam format escaped
                self.entry_furyo_no.get() or "",  # 不良発生№ (col 12)
            ]
            ws.append(vals)
            # reindex setelah append untuk jaga konsistensi
            self.reindex_excel(ws)
            wb.save(self.excel_path)
            self.load_excel_to_tree()
            messagebox.showinfo(JP_LABELS["success"], JP_LABELS["added_ok"])

        except Exception as e:
            messagebox.showerror(
                JP_LABELS["error"], f"{JP_LABELS['error_add_row']} {str(e)}"
            )

    def on_tree_select(self, event):
        sel = self.tree.selection()
        if not sel:
            self.selected_row = None
            self.update_button_states()
            return

        try:
            vals = self.tree.item(sel[0], "values")
            self.selected_row = (
                self.tree.index(sel[0]) + self.data_start_row
            )  # Excel row index
            self.fill_form_with_data(vals)
            self.update_button_states()  # Update button states when row is selected

        except Exception as e:
            messagebox.showerror(
                JP_LABELS["error"], f"{JP_LABELS['error_add_row']} {str(e)}"
            )

    def update_row(self):
        if not self.excel_path or not self.selected_sheet or not self.selected_row:
            messagebox.showwarning(JP_LABELS["warning"], JP_LABELS["pick_excel_first"])
            return

        try:
            wb = load_workbook(self.excel_path)
            ws = wb[self.selected_sheet]
            row = self.selected_row

            # Escape path untuk kompatibilitas Jepang
            escaped_renrakusho_path = escape_path_for_japanese_locale(
                self.entry_renrakusho.get() or ""
            )

            ws.cell(row=row, column=1).value = self.entry_hassei_month.get() or ""
            # col 2 (累計) akan direindex ulang
            ws.cell(row=row, column=3).value = self.entry_no.get() or ""
            ws.cell(row=row, column=4).value = self.entry_date.get() or ""
            ws.cell(row=row, column=5).value = self.cbo_koumoku.get() or ""
            ws.cell(row=row, column=6).value = self.entry_jishou.get() or ""
            ws.cell(row=row, column=7).value = self.cbo_ichiji.get() or ""
            ws.cell(row=row, column=8).value = self.cbo_niji.get() or ""
            ws.cell(row=row, column=9).value = self.entry_hinban.get() or ""
            ws.cell(row=row, column=10).value = self.cbo_supplier.get() or ""
            ws.cell(row=row, column=11).value = (
                escaped_renrakusho_path  # Path dalam format escaped
            )
            ws.cell(row=row, column=12).value = self.entry_furyo_no.get() or ""

            # reindex all
            self.reindex_excel(ws)
            wb.save(self.excel_path)
            self.load_excel_to_tree()
            messagebox.showinfo(JP_LABELS["success"], JP_LABELS["updated_ok"])

        except Exception as e:
            messagebox.showerror(
                JP_LABELS["error"], f"{JP_LABELS['error_select_row']} {str(e)}"
            )

    def delete_row(self):
        if not self.excel_path or not self.selected_sheet or not self.selected_row:
            messagebox.showwarning(JP_LABELS["warning"], JP_LABELS["pick_excel_first"])
            return

        if not messagebox.askyesno(JP_LABELS["confirm"], JP_LABELS["confirm_delete"]):
            return

        try:
            wb = load_workbook(self.excel_path)
            ws = wb[self.selected_sheet]
            ws.delete_rows(self.selected_row, 1)
            self.reindex_excel(ws)
            wb.save(self.excel_path)
            self.load_excel_to_tree()
            self.clear_form()
            messagebox.showinfo(JP_LABELS["deleted"], JP_LABELS["deleted_ok"])

        except Exception as e:
            messagebox.showerror(
                JP_LABELS["error"], f"{JP_LABELS['error_delete_row']} {str(e)}"
            )

    def reindex_excel(self, ws):
        """Set ulang 累計 (col 2) dan № (col 3) agar berurutan mulai 1"""
        for i, row in enumerate(
            ws.iter_rows(min_row=self.data_start_row, values_only=False), start=1
        ):
            if len(row) > 1:
                row[1].value = i  # 累計 (B)
            if len(row) > 2:
                row[2].value = i  # № (C)
